

'''
接口自动化测试步骤：
1:excel 测试用例的准备ok,代码自动去读取测试数据
2：发送接口请求，得到响应结果
3：断言：执行结果vs预期结果   --通过、不通过
4：写入最终结果到excel表格（不涉及输出测试报告）

#怎么进行excel操作
# 操作excel 的第三方库 ： openpyxl库 ， 1）安装：pip install openpyxl 2)导入 import openpyxl
excel文档三大对象：
WorkBook  工作簿
Sheet     表单
Cell      单元格
excel文件相关操作
wb = load_workbook('sample.xlsx')   #打开一个已经存在的excel 文件
rows = sh.max_row    #获取总行数
columns = sh.max_column    #获取总列数
cl = sh.cell(row=1,column = 1).value    #获取对应行列的单元格值
sh.cell(row=1,column=1).value = 'hello'   #修改某个指定的单元格值
wb.save("hello.xlsx")    #保存workbook，这样才可以保存刚刚修改写入的值（文档状态为关闭）
注意：openpyxl中行和列的起始标识是1


'''
# import openpyxl
# 只需要用到这个库的load_workbook 函数
from openpyxl import load_workbook

# wb = load_workbook('testcase_api_wuye.xlsx')  # 打开excel文件
# sh = wb['register']
def read_case(filename,sheetname):
    wb = load_workbook(filename)     #打开excel文件
    sh = wb[sheetname]  #打开某个表单
    # cl = sh.cell(row=1,column=1).value  #通过表单获取对应的行，列值
    # print(cl)
    max_row = sh.max_row    #获取总行数
    case_list = []      #新建一个新列表
    for i in range(2,max_row+1):
        case_dict= dict(
        case_id = sh.cell( row=i, column = 1),  #获取case_id
        url = sh.cell(row=i,column= 5).value,   # 获取url值
        data = sh.cell(row=i,column= 6).value,    #获取data
        expect = sh.cell(row=i,column=7).value )   #获取expect值
        case_list.append(case_dict)
    print(case_list)
    return case_list
#调用读取函数
case_list=read_case('testcase_api_wuye.xlsx','register')

#写入
# wb = load_workbook('testcase_api_wuye.xlsx')     #打开excel文件
# sh = wb['register']   #打开某个表单
# sh.cell(row=2,column=8).value = 'passed'    #写入数据
# wb.save('testcase_api_wuye.xlsx')

#将写入封装成函数
def write_result(filename,sheetname,row,column,final_result):
    wb= load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value= final_result    #写入数据
    wb.save(filename)  #一定要保存



