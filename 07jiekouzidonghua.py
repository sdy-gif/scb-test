# 接口自动化测试步骤：
# # 1:excel 测试用例的准备ok,代码自动去读取测试数据
# # 2：发送接口请求，得到响应结果
# # 3：断言：执行结果vs预期结果   --通过、不通过
# # 4：写入最终结果到excel表格（不涉及输出测试报告）
from openpyxl import load_workbook
import requests

#读取excel表函数
def read_case(filename,sheetname):
    wb = load_workbook(filename)     #打开excel文件
    sh = wb[sheetname]  #打开某个表单
    # cl = sh.cell(row=1,column=1).value  #通过表单获取对应的行，列值
    # print(cl)
    max_row = sh.max_row    #获取总行数
    case_list = []      #新建一个新列表
    for i in range(2,max_row+1):
        case_dict= dict(
        case_id = sh.cell( row=i, column = 1),  #获取case_id
        url = sh.cell(row=i,column= 5).value,   # 获取url值
        data = sh.cell(row=i,column= 6).value,    #获取data
        expect = sh.cell(row=i,column=7).value )   #获取expect值
        case_list.append(case_dict)
    print(case_list)
    return case_list

# read_case('testcase_api_wuye.xlsx','register')

# 发送post请求函数
def api_fun (url,data):
    header = {'X-Lemonban-Media-Type': 'lemonban.v2',
                    'Content-Type': 'application/json'}
    res = requests.post(url=url,json=data,headers=header).json()
    return res

#将写入封装成函数
def write_result(filename,sheetname,row,column,final_result):
    wb= load_workbook(filename)
    sh = wb[sheetname]
    sh.cell(row=row,column=column).value= final_result    # 写入数据
    wb.save(filename)  # 一定要保存

# eval()函数   --运行被字符串包裹的python表达式
# print(eval('3+2'))  #去掉字符串的引号之后，运行3+2  所以输出5
# data = '{"phone":"13112340000","pwd":"1234567a","rePwd":"1234567a","userName":"柠檬dls","verificationCode":"lemon"}'
# print(eval(data))   #去掉字符串之后 ，运行里面的内容，因为是字典的格式，所以出来的是字典

# 完整的接口自动化测试，封装成函数
def excute_fun(filename ,sheetname):
    cases = read_case(filename, sheetname)
    for case in cases:
        cases_id = case['case_id']
        url = case['url']
        data = eval(case['data'])
        expect = eval(case['expect'])
        # print(case)
        # print(cases_id,url,type(data),type(expect))  # 这里的data 和 expect 读取出来的是字符串类型

        # 获取期望code msg
        expect_code = expect['code']
        expect_msg = expect['msg']

        real_result = api_fun(url, data)  # 得到响应结果
        # print (real_result)

        # 获取实际运行得到的code ,msg
        real_code = real_result['code']
        real_msg = real_result['msg']

        # 断言 ：比较期望的code ,msg和 实际的code .msg
        if expect_code == real_code and expect_msg == real_msg:
            print('第{}条测试用例通过'.format(cases_id))
            final_re = 'passed'
        else:
            print('第{}条测试用例执行不通过'.format(cases_id))
            final_re = 'falied'

        # 写入数据
        write_result(filename, sheetname, cases_id + 1, 8, final_re)